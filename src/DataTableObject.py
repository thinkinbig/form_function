import logging

import pandas as pd
from openpyxl.reader.excel import load_workbook

from src.utils import EmptyException, ignore_value
from openpyxl.styles import Font, PatternFill, Color

# 判断是否需要高亮
HIGHLIGHT_SHIFT = 0
HIGHLIGHT_MARK = 1 << HIGHLIGHT_SHIFT
# 判断是否需要变色
FONT_SHIFT = 1
FONT_MARK = 1 << FONT_SHIFT
NO_MARK = 0x0


class CustomExcelPainter:
    YELLOW = Color(rgb='00FFFF00')
    RED = Color(rgb='00FF0000')

    def __init__(self, output_path: str):
        self.pattern_fill = PatternFill(start_color=self.YELLOW, end_color=self.YELLOW, fill_type='solid')
        self.font = Font(color=self.RED)
        self.output_path = output_path
        self.workbook = load_workbook(output_path)
        self.worksheet = self.workbook.active

    def paint(self, row_index: int, column_index: int, pattern=False, font=False):
        """
        Paint the cell with yellow color and red font
        :param row_index: row index
        :param column_index: column index
        :param pattern: True if need to paint yellow
        :param font: True if need to paint red
        :return: None
        """
        cell = self.worksheet.cell(row=row_index, column=column_index)
        if pattern:
            cell.fill = self.pattern_fill
        if font:
            cell.font = self.font

    def mark_row(self, row_index: int, color: Color):
        """
        Mark the row with color
        :param row_index: row index
        :param color: color
        :return: None
        """
        for cell in self.worksheet[row_index]:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

    def save(self):
        self.workbook.save(self.output_path)

    def column_index_by_name(self, column_name):
        for cell in next(self.worksheet.rows):
            if cell.value == column_name:
                return cell.column
        raise EmptyException(f"Column {column_name} not found")


class DataTableObject:

    def __init__(self, standard_path: str, input_path: str):
        self.standard_path = standard_path
        self.input_path = input_path
        self.current_row = 1
        self._init_dataframe()

    def __repr__(self):
        return self.dataframe.__repr__()

    def __getitem__(self, item):
        entity = self.dataframe.iloc[self.current_row][item]
        if pd.isnull(entity):
            raise EmptyException(f"Row {self.current_row} of {item} not found")
        return entity

    def __setitem__(self, key, value):
        self.dataframe.loc[self.current_row, key] = value
        logging.debug(f"Row {self.current_row} set {key} to {value}")

    def next(self):
        self.current_row += 1

    @property
    def eof(self) -> bool:
        """
        Check if end of file
        Returns
        -------
        bool : True if end of file, False otherwise
        """
        return self.current_row >= self.dataframe.shape[0]

    def export_dataframe(self, output_path: str):
        """
        Export dataframes to excel file
        :param output_path: output path
        :return: None
        """
        self.dataframe.to_excel(output_path, index=False, engine='openpyxl')
        painter = CustomExcelPainter(output_path)
        self._paint_cv_cells(painter)
        self._paint_open_rate_cells(painter)
        self._paint_entire_line(painter)
        self._remove_flag_columns(painter)
        painter.save()

    def _init_dataframe(self):
        """
        Read important table columns from Excel file into model
        """
        standard_dataframe = pd.read_excel(self.standard_path, index_col=None)
        if standard_dataframe.shape[1] < 2:
            raise Exception(f"Standard file {self.standard_path} must have 2 columns")
        # get the first two columns
        standard_dataframe = standard_dataframe.iloc[:, :2]
        self.dataframe = standard_dataframe.transpose().reset_index(drop=True)
        self.dataframe.columns = self.dataframe.iloc[0]
        self.dataframe = self.dataframe.drop(self.dataframe.index[0])
        self.header = self.dataframe.columns.values.tolist()

        # fill input data to dataframe
        input_df = pd.read_excel(self.input_path, index_col=None)
        # drop the first row which is the chinese header
        input_df = input_df.iloc[1:]
        self.dataframe = pd.concat([self.dataframe, input_df], axis=0, ignore_index=True, sort=False)

    @staticmethod
    @ignore_value
    def _paint_cv_cells(painter: CustomExcelPainter):
        cv_flag = ['F_BEM_QtyJSCVPD1', 'F_BEM_QtyJSCVPD2', 'F_BEM_QtyJSCVPD3']
        cv_columns = ['F_BEM_QtyJSCV', 'F_BEM_QtyJSCVnro', 'F_BEM_QtyJSCVmax']
        for row_index in range(3, painter.worksheet.max_row + 1):
            for i in range(3):
                flag = painter.worksheet.cell(row=row_index, column=painter.column_index_by_name(cv_flag[i])).value
                assert flag is not None
                column = painter.column_index_by_name(cv_columns[i])
                painter.paint(row_index, column, font=(flag & FONT_MARK) != 0, pattern=(flag & HIGHLIGHT_MARK) != 0)

    @staticmethod
    @ignore_value
    def _paint_open_rate_cells(painter: CustomExcelPainter):
        open_rate_flag = ['F_BEM_JSKDPD1', 'F_BEM_JSKDPD2', 'F_BEM_JSKDPD3']
        open_rate_columns = ['F_BEM_JSKDMIN', 'F_BEM_JSKDNOR', 'F_BEM_JSKDMAX']
        for row_index in range(3, painter.worksheet.max_row + 1):
            for i in range(3):
                flag = painter.worksheet.cell(row=row_index,
                                              column=painter.column_index_by_name(open_rate_flag[i])).value
                assert flag is not None
                column = painter.column_index_by_name(open_rate_columns[i])
                painter.paint(row_index, column, font=(flag & FONT_MARK) != 0, pattern=(flag & HIGHLIGHT_MARK) != 0)

    @staticmethod
    @ignore_value
    def _paint_entire_line(painter: CustomExcelPainter):
        for row_index in range(3, painter.worksheet.max_row + 1):
            column_flag = "F_BEM_ROW_ERROR"
            flag = painter.worksheet.cell(row=row_index, column=painter.column_index_by_name(column_flag)).value
            logging.debug(f"f_BEM_ROW_ERROR {flag}")
            if flag is not None:
                painter.mark_row(row_index, CustomExcelPainter.RED)

    @staticmethod
    @ignore_value
    def _remove_flag_columns(painter: CustomExcelPainter):
        flag_columns = ['F_BEM_QtyJSCVPD1', 'F_BEM_QtyJSCVPD2', 'F_BEM_QtyJSCVPD3',
                        'F_BEM_JSKDPD1', 'F_BEM_JSKDPD2', 'F_BEM_JSKDPD3']
        for column in flag_columns:
            painter.worksheet.delete_cols(painter.column_index_by_name(column))
